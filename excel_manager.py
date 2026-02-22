import openpyxl
from datetime import datetime
import os

ARCHIVO = "servidores.xlsx"

def inicializar_excel():
    if not os.path.exists(ARCHIVO):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Servidores"
        ws.append(["ID", "Nombre", "IP", "Versión", "Tipo", "Estado", "Fecha Registro"])
        wb.save(ARCHIVO)

def obtener_nuevo_id():
    wb = openpyxl.load_workbook(ARCHIVO)
    ws = wb.active
    filas = list(ws.iter_rows(values_only=True))[1:]
    return len(filas) + 1

def agregar_servidor(nombre, ip, version, tipo):
    wb = openpyxl.load_workbook(ARCHIVO)
    ws = wb.active
    nuevo_id = obtener_nuevo_id()
    fecha = datetime.now().strftime("%d/%m/%Y %H:%M")
    ws.append([nuevo_id, nombre, ip, version, tipo, "🔴 Offline", fecha])
    wb.save(ARCHIVO)
    return nuevo_id

def obtener_servidores():
    wb = openpyxl.load_workbook(ARCHIVO)
    ws = wb.active
    return list(ws.iter_rows(values_only=True))[1:]

def actualizar_estado(server_id, nuevo_estado):
    wb = openpyxl.load_workbook(ARCHIVO)
    ws = wb.active
    for fila in ws.iter_rows(min_row=2):
        if fila[0].value == server_id:
            fila[5].value = nuevo_estado
            wb.save(ARCHIVO)
            return True
    return False

def obtener_estadisticas():
    servidores = obtener_servidores()
    total = len(servidores)
    online = sum(1 for s in servidores if "Online" in str(s[5]))
    offline = total - online
    versiones = {}
    for s in servidores:
        v = s[3]
        versiones[v] = versiones.get(v, 0) + 1
    return total, online, offline, versiones